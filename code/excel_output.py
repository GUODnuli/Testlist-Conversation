import openpyxl
import os
import shutil
from .ListNode import TreeNode

def process_to_excel(root: TreeNode, output_file: str, template_file: str, config_data: dict):
    # 检查excel文件是否存在，若不存在则创建一个
    if not os.path.exists(output_file):
        # 确保 template.xlsx 文件存在于预期的位置，并确保有权限读取和写入这些文件。
        shutil.copy(template_file, output_file)
    
    workbook = openpyxl.load_workbook(output_file)
    sheet = workbook.active
    required_fields = config_data["必填项"]
    sep = config_data["用例目录识别符"]
    
    root_folder = required_fields["根目录"]
    case_type = required_fields["用例类型"]
    cases_tatus = required_fields["用例状态"]
    creator = required_fields["创建人"]

    # 设置列标题
    # headers = ["用例目录", "用例名称", "需求ID", "前置条件", "用例步骤", "预期结果", "用例类型", "用例状态", "创建人"]
    # for col_num, header in enumerate(headers, 1):
    #     sheet.cell(row=1, column=col_num).value = header
    
    sheet.column_dimensions["A"].width = 70
    # sheet.column_dimensions["B"].width = 35
    # sheet.column_dimensions["C"].width = 100
    # sheet.column_dimensions["D"].width = 150

    # 用于存储当前路径的列表
    current_path = []
    row_num = 1
    
    def traverse_and_write(node: TreeNode, depth=0):
        nonlocal current_path  # 使用外部函数的变量
        nonlocal row_num
        
        # 构建路径
        if node.title.startswith(sep):
            current_path.append(node.title)
        else:
            row_num = row_num + 1  # 获取当前应写入的行号
            case_path = root_folder+"".join(current_path)
            sheet.cell(row=row_num, column=1).value = case_path
            sheet.cell(row=row_num, column=2).value = node.title
            sheet.cell(row=row_num, column=3).value = None
            sheet.cell(row=row_num, column=4).value = "无"
            sheet.cell(row=row_num, column=5).value = node.step
            sheet.cell(row=row_num, column=6).value = node.result
            sheet.cell(row=row_num, column=7).value = case_type
            sheet.cell(row=row_num, column=8).value = cases_tatus
            sheet.cell(row=row_num, column=9).value = node.level
            sheet.cell(row=row_num, column=10).value = creator
        
        # 遍历子节点
        for child in node.children:
            traverse_and_write(child, depth + 1)
        
        # 当完成当前节点及其子节点的处理后，从路径中删除该节点（如果它是一个文件夹）
        if node.title.startswith(sep):
            current_path.pop()

    traverse_and_write(root)
    workbook.save(output_file)