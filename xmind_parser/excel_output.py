import openpyxl
from .Data_struct import TreeNode

def process_to_excel(root: TreeNode, output_file: str, config_data: dict):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    required_fields = config_data["必填项"]
    sep = config_data["用例目录识别符"]
    
    root_folder = required_fields["根目录"]
    case_type = required_fields["用例类型"]
    cases_tatus = required_fields["用例状态"]
    creator = required_fields["创建人"]

    # 设置列标题
    headers = ["完整用例文件夹路径", "用例标题", "用例步骤", "用例结果", "用例类型", "用例状态", "创建人"]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header
    
    sheet.column_dimensions["A"].width = 70
    sheet.column_dimensions["B"].width = 35
    sheet.column_dimensions["C"].width = 100
    sheet.column_dimensions["D"].width = 150

    # 用于存储当前路径的列表
    current_path = []
    
    def traverse_and_write(node: TreeNode, depth=0):
        nonlocal current_path  # 使用外部函数的变量
        
        # 构建路径
        if node.title.startswith(sep):
            current_path.append(node.title)
        else:
            row_num = sheet.max_row + 1  # 获取当前应写入的行号
            sheet.cell(row=row_num, column=1).value = root_folder+sep+"".join(current_path)
            sheet.cell(row=row_num, column=2).value = node.title
            sheet.cell(row=row_num, column=3).value = node.step
            sheet.cell(row=row_num, column=4).value = node.result
            sheet.cell(row=row_num, column=5).value = case_type
            sheet.cell(row=row_num, column=6).value = cases_tatus
            sheet.cell(row=row_num, column=7).value = creator
        
        # 遍历子节点
        for child in node.children:
            traverse_and_write(child, depth + 1)
        
        # 当完成当前节点及其子节点的处理后，从路径中删除该节点（如果它是一个文件夹）
        if node.title.startswith(sep):
            current_path.pop()

    traverse_and_write(root)
    workbook.save(output_file)